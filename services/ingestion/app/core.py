from __future__ import annotations

import csv
import hashlib
from pathlib import Path

import fitz
import openpyxl

MAX_CHARS = 1100
OVERLAP_CHARS = 160


def merge_page_blocks(page: fitz.Page) -> list[dict[str, object]]:
    """Create deterministic, page-local citation regions from visual PDF blocks."""
    blocks = sorted(page.get_text("blocks"), key=lambda block: (round(block[1], 2), round(block[0], 2)))
    regions: list[dict[str, object]] = []
    buffer: list[str] = []
    x0 = y0 = float("inf")
    x1 = y1 = float("-inf")

    def flush() -> None:
        nonlocal buffer, x0, y0, x1, y1
        text = "\n".join(buffer).strip()
        if text:
            regions.append({
                "page_number": page.number + 1,
                "bbox": [x0, y0, x1, y1],
                "text": text,
                "content_hash": hashlib.sha256(text.encode("utf-8")).hexdigest(),
            })
        buffer, x0, y0, x1, y1 = [], float("inf"), float("inf"), float("-inf"), float("-inf")

    for block in blocks:
        block_text = " ".join(block[4].split())
        if not block_text:
            continue
        if buffer and sum(len(item) for item in buffer) + len(block_text) > MAX_CHARS:
            overlap = buffer[-1][-OVERLAP_CHARS:]
            flush()
            if overlap:
                buffer.append(overlap)
        buffer.append(block_text)
        x0, y0 = min(x0, block[0]), min(y0, block[1])
        x1, y1 = max(x1, block[2]), max(y1, block[3])
    flush()
    return regions


def parse_pdf(path: Path) -> list[dict[str, object]]:
    try:
        with fitz.open(path) as document:
            return [region for page in document for region in merge_page_blocks(page)]
    except fitz.FileDataError as error:
        raise ValueError("Unreadable PDF") from error


def render_row_text(headers: list[str] | None, values: list[object], *, prefix: str = "") -> str:
    """Render a single tabular row as a human-readable 'field: value' citation string."""
    fields: list[str] = []
    for index, raw_value in enumerate(values):
        value = "" if raw_value is None else str(raw_value).strip()
        if not value:
            continue
        label = headers[index].strip() if headers and index < len(headers) and headers[index] and str(headers[index]).strip() else f"field{index + 1}"
        fields.append(f"{label}: {value}")
    body = " | ".join(fields)
    return f"{prefix}{body}" if prefix else body


def parse_csv(path: Path) -> list[dict[str, object]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            rows = list(reader)
    except (UnicodeDecodeError, csv.Error) as error:
        raise ValueError("Unreadable CSV") from error

    if not rows:
        return []

    headers = [str(cell) for cell in rows[0]]
    regions: list[dict[str, object]] = []
    for row_index, row in enumerate(rows[1:], start=1):
        if not any(str(cell).strip() for cell in row):
            continue
        text = render_row_text(headers, list(row))
        if not text:
            continue
        regions.append({
            "page_number": row_index,
            "bbox": None,
            "text": text,
            "content_hash": hashlib.sha256(text.encode("utf-8")).hexdigest(),
        })
    return regions


def parse_xlsx(path: Path) -> list[dict[str, object]]:
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as error:  # openpyxl raises assorted zip/xml errors on corrupt files
        raise ValueError("Unreadable XLSX") from error

    try:
        sheet_names = workbook.sheetnames
        multi_sheet = len(sheet_names) > 1
        regions: list[dict[str, object]] = []
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            headers: list[str] | None = None
            data_row_number = 0
            for row in sheet.iter_rows(values_only=True):
                values = list(row)
                if headers is None:
                    headers = [str(cell) if cell is not None else "" for cell in values]
                    continue
                if not any(cell is not None and str(cell).strip() for cell in values):
                    continue
                data_row_number += 1
                prefix = f"{sheet_name} row {data_row_number} — " if multi_sheet else ""
                text = render_row_text(headers, values, prefix=prefix)
                if not text:
                    continue
                regions.append({
                    "page_number": data_row_number,
                    "bbox": None,
                    "text": text,
                    "content_hash": hashlib.sha256(text.encode("utf-8")).hexdigest(),
                })
        return regions
    except ValueError:
        raise
    except Exception as error:
        raise ValueError("Unreadable XLSX") from error
    finally:
        workbook.close()
